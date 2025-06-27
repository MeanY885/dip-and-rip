# app.py (with data viewer and price monitor)
from flask import Flask, render_template, request, jsonify, redirect
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import desc
import pandas as pd
import numpy as np
import requests
import plotly.graph_objects as go
import plotly.utils
from plotly.subplots import make_subplots
import json
import os
from datetime import datetime, timedelta
import logging
import itertools
import shutil
from flask import send_file
import tempfile
import hashlib
import re
import csv
import io
from werkzeug.utils import secure_filename
try:
    import pdfplumber  # pdfplumber for PDF parsing
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    logger.warning("pdfplumber not installed. PDF parsing will be disabled.")
import xlrd
import openpyxl
from difflib import SequenceMatcher

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# File upload configuration
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = 'temp_uploads'
ALLOWED_EXTENSIONS = {'pdf'}

# Database configuration
database_url = os.environ.get('SQLALCHEMY_DATABASE_URI', 'sqlite:///finance_tracker.db')
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Handle database directory creation for file-based SQLite
if database_url.startswith('sqlite:///'):
    # Extract the path (handle both relative and absolute paths)
    if database_url.startswith('sqlite:////'):
        # Absolute path (4 slashes)
        db_path = database_url[10:]  # Remove 'sqlite:///'
    else:
        # Relative path (3 slashes)
        db_path = database_url[10:]  # Remove 'sqlite:///'
    
    db_dir = os.path.dirname(db_path)
    
    if db_dir:
        try:
            os.makedirs(db_dir, mode=0o755, exist_ok=True)
            logger.info(f"Database directory ensured: {db_dir}")
            
            # Verify we can write to the directory
            if os.access(db_dir, os.W_OK):
                logger.info(f"Write access confirmed for: {db_dir}")
            else:
                logger.warning(f"No write access to: {db_dir}")
                
        except Exception as e:
            logger.error(f"Failed to create database directory {db_dir}: {e}")
            # Fallback to current directory
            app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///finance_tracker.db'
            logger.info("Falling back to current directory for database")

db = SQLAlchemy(app)

# Helper function for safe date formatting
def format_date_safe(date_obj):
    """Safely format date object to ISO format string"""
    if date_obj is None:
        return None
    
    # If it's already a string, return as is
    if isinstance(date_obj, str):
        return date_obj
    
    # If it has isoformat method (datetime.date or datetime.datetime)
    if hasattr(date_obj, 'isoformat'):
        return date_obj.isoformat()
    
    # Fallback to string conversion
    return str(date_obj)

# Database Models
class Investment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    start_investment = db.Column(db.Float, nullable=False)
    current_value = db.Column(db.Float, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    yearly_records = db.relationship('YearlyRecord', backref='investment_ref', lazy=True, cascade='all, delete-orphan')
    monthly_records = db.relationship('MonthlyRecord', backref='investment_ref', lazy=True, cascade='all, delete-orphan')

class YearlyRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    investment_id = db.Column(db.Integer, db.ForeignKey('investment.id'), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    value = db.Column(db.Float, nullable=False)
    notes = db.Column(db.Text)
    date = db.Column(db.Date, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    __table_args__ = (db.UniqueConstraint('investment_id', 'year', name='unique_investment_year'),)

class MonthlyRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    investment_id = db.Column(db.Integer, db.ForeignKey('investment.id'), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    value = db.Column(db.Float, nullable=False)
    notes = db.Column(db.Text)
    date = db.Column(db.Date, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    __table_args__ = (db.UniqueConstraint('investment_id', 'year', 'month', name='unique_investment_year_month'),)

class UserPreferences(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    widget_order = db.Column(db.Text)  # JSON string of widget order
    investment_order = db.Column(db.Text)  # JSON string of investment order
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class BitcoinTrade(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    status = db.Column(db.String(20), nullable=False)  # 'Open', 'Closed'
    date = db.Column(db.Date, nullable=False)
    type = db.Column(db.String(10), nullable=False, default='BTC')  # 'BTC'
    initial_investment_gbp = db.Column(db.Float, nullable=False)  # Renamed from gbp_investment
    btc_buy_price = db.Column(db.Float, nullable=False)  # Renamed from btc_value
    btc_sell_price = db.Column(db.Float, nullable=True)  # Renamed from crypto_value, nullable for open trades
    profit = db.Column(db.Float, nullable=True)  # Auto-calculated, nullable for open trades
    fee = db.Column(db.Float, nullable=True)  # Nullable for open trades
    btc_amount = db.Column(db.Float, nullable=True)  # Store the amount of BTC purchased
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class BitcoinPriceHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, nullable=False, unique=True)
    price_gbp = db.Column(db.Float, nullable=False)
    price_usd = db.Column(db.Float, nullable=True)  # Optional USD price for reference
    volume = db.Column(db.Float, nullable=True)  # Trading volume if available
    source = db.Column(db.String(20), default='kraken')  # Data source
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    __table_args__ = (db.Index('idx_timestamp', 'timestamp'),)

class FinanceCategory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    color = db.Column(db.String(7), default='#007bff')  # Hex color code
    description = db.Column(db.Text, nullable=True)
    budget = db.Column(db.Float, nullable=True)  # Monthly budget for this category
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    transactions = db.relationship('FinanceTransaction', backref='category_ref', lazy=True)

class FinanceTransaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    amount = db.Column(db.Float, nullable=False)
    description = db.Column(db.Text, nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('finance_category.id'), nullable=True)
    month = db.Column(db.Integer, nullable=False)  # 1-12
    year = db.Column(db.Integer, nullable=False)
    source_file = db.Column(db.String(255), nullable=True)  # Track which file imported from
    source_row = db.Column(db.Text, nullable=True)  # Original row data for duplicate detection
    hash_value = db.Column(db.String(64), nullable=True, index=True)  # For duplicate detection
    is_duplicate = db.Column(db.Boolean, default=False)
    confidence_score = db.Column(db.Float, nullable=True)  # Auto-categorization confidence
    is_recurring = db.Column(db.Boolean, default=False)  # Mark if this is a recurring transaction
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    __table_args__ = (db.Index('idx_date_amount', 'date', 'amount'),
                      db.Index('idx_year_month', 'year', 'month'))

class FinanceCategoryLearning(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description_pattern = db.Column(db.Text, nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('finance_category.id'), nullable=False)
    frequency = db.Column(db.Integer, default=1)  # How often this pattern matches this category
    confidence = db.Column(db.Float, default=1.0)  # 0-1 confidence score
    last_used = db.Column(db.DateTime, default=datetime.utcnow)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    category = db.relationship('FinanceCategory', backref='learning_patterns', lazy=True)

# Create tables
with app.app_context():
    db.create_all()
    
    # Handle migration for BitcoinTrade table if needed
    try:
        # Check if we need to migrate old column names to new ones
        from sqlalchemy import inspect, text
        
        inspector = inspect(db.engine)
        if 'bitcoin_trade' in inspector.get_table_names():
            columns = [col['name'] for col in inspector.get_columns('bitcoin_trade')]
            
            # Check if old columns exist
            if 'gbp_investment' in columns and 'initial_investment_gbp' not in columns:
                logger.info("Migrating BitcoinTrade table to new schema...")
                
                # Create migration SQL
                migration_sql = """
                    ALTER TABLE bitcoin_trade ADD COLUMN initial_investment_gbp FLOAT;
                    ALTER TABLE bitcoin_trade ADD COLUMN btc_buy_price FLOAT;
                    ALTER TABLE bitcoin_trade ADD COLUMN btc_sell_price FLOAT;
                    ALTER TABLE bitcoin_trade ADD COLUMN btc_amount FLOAT;
                    
                    UPDATE bitcoin_trade 
                    SET initial_investment_gbp = gbp_investment,
                        btc_buy_price = btc_value,
                        btc_sell_price = crypto_value,
                        btc_amount = gbp_investment / btc_value;
                    
                    UPDATE bitcoin_trade
                    SET status = 'Closed'
                    WHERE btc_sell_price IS NOT NULL AND fee IS NOT NULL;
                    
                    UPDATE bitcoin_trade
                    SET status = 'Open'
                    WHERE btc_sell_price IS NULL OR fee IS NULL;
                """
                
                # Execute migration
                with db.engine.connect() as conn:
                    for statement in migration_sql.strip().split(';'):
                        if statement.strip():
                            conn.execute(text(statement))
                    conn.commit()
                
                logger.info("Migration completed successfully")
                
                # Drop old columns after successful migration
                drop_sql = """
                    ALTER TABLE bitcoin_trade DROP COLUMN gbp_investment;
                    ALTER TABLE bitcoin_trade DROP COLUMN btc_value;
                    ALTER TABLE bitcoin_trade DROP COLUMN crypto_value;
                """
                
                with db.engine.connect() as conn:
                    for statement in drop_sql.strip().split(';'):
                        if statement.strip():
                            try:
                                conn.execute(text(statement))
                            except Exception as e:
                                logger.warning(f"Could not drop old column: {e}")
                    conn.commit()
                    
    except Exception as e:
        logger.warning(f"Migration check failed: {e}")
        # If migration fails, just continue - the table might be new

class BTCBacktester:
    def __init__(self, lookback_days=365, investment_value=1000, buy_dip_percent=5, sell_gain_percent=10, transaction_fee_percent=0.1):
        self.lookback_days = lookback_days
        self.investment_value = investment_value
        self.buy_dip_percent = buy_dip_percent / 100
        self.sell_gain_percent = sell_gain_percent / 100
        self.transaction_fee_percent = transaction_fee_percent / 100
        self.data = None
        self.trades = []
        
    def fetch_data(self):
        """Fetch BTC data from Kraken API"""
        try:
            url = "https://api.kraken.com/0/public/OHLC"
            end_date = datetime.now()
            start_date = end_date - timedelta(days=self.lookback_days + 30)
            since = int(start_date.timestamp())
            
            logger.info(f"Fetching Kraken data from {start_date} to {end_date}")
            
            pairs_to_try = [
                ("XXBTZGBP", "BTC-GBP"),
                ("XXBTZUSD", "BTC-USD")
            ]
            
            for kraken_pair, display_name in pairs_to_try:
                try:
                    logger.info(f"Trying Kraken pair: {kraken_pair} ({display_name})")
                    
                    params = {
                        'pair': kraken_pair,
                        'interval': 1440,
                        'since': since
                    }
                    
                    response = requests.get(url, params=params, timeout=10)
                    response.raise_for_status()
                    
                    data = response.json()
                    
                    if 'error' in data and data['error']:
                        logger.error(f"Kraken API error for {kraken_pair}: {data['error']}")
                        continue
                    
                    if 'result' not in data or not data['result']:
                        logger.warning(f"No result data for {kraken_pair}")
                        continue
                    
                    pair_data = None
                    for key in data['result'].keys():
                        if key != 'last':
                            pair_data = data['result'][key]
                            break
                    
                    if not pair_data or len(pair_data) < 10:
                        logger.warning(f"Insufficient data for {kraken_pair}")
                        continue
                    
                    df_data = []
                    for row in pair_data:
                        df_data.append({
                            'timestamp': int(row[0]),
                            'Open': float(row[1]),
                            'High': float(row[2]),
                            'Low': float(row[3]),
                            'Close': float(row[4]),
                            'Volume': float(row[6])
                        })
                    
                    df = pd.DataFrame(df_data)
                    df['Date'] = pd.to_datetime(df['timestamp'], unit='s')
                    df.set_index('Date', inplace=True)
                    df.drop('timestamp', axis=1, inplace=True)
                    df = df.sort_index()
                    
                    logger.info(f"Successfully fetched {len(df)} rows of data for {display_name}")
                    self.data = df
                    return True
                    
                except Exception as e:
                    logger.error(f"Error processing {kraken_pair}: {str(e)}")
                    continue
            
            logger.error("All Kraken pairs failed to fetch data")
            return False
            
        except Exception as e:
            logger.error(f"General error fetching Kraken data: {str(e)}")
            return False
    
    def run_backtest_with_params(self, buy_dip_percent, sell_gain_percent):
        """Run backtest with specific parameters (for optimization)"""
        if self.data is None or self.data.empty:
            return None
            
        # Calculate signals with given parameters
        buy_dip_decimal = buy_dip_percent / 100
        sell_gain_decimal = sell_gain_percent / 100
        
        returns = self.data['Close'].pct_change()
        buy_signals = returns <= -buy_dip_decimal
        sell_signals = returns >= sell_gain_decimal
        
        # Run simulation
        cash = self.investment_value
        btc_holdings = 0
        holding_btc = False
        total_fees_paid = 0
        num_trades = 0
        portfolio_values = []
        
        for i, (date, row) in enumerate(self.data.iterrows()):
            price = row['Close']
            
            # Buy signal
            if buy_signals.iloc[i] and not holding_btc and cash > 0:
                transaction_fee = cash * self.transaction_fee_percent
                cash_after_fees = cash - transaction_fee
                btc_holdings = cash_after_fees / price
                total_fees_paid += transaction_fee
                cash = 0
                holding_btc = True
                num_trades += 1
            
            # Sell signal
            elif sell_signals.iloc[i] and holding_btc and btc_holdings > 0:
                cash_before_fees = btc_holdings * price
                transaction_fee = cash_before_fees * self.transaction_fee_percent
                cash = cash_before_fees - transaction_fee
                total_fees_paid += transaction_fee
                btc_holdings = 0
                holding_btc = False
                num_trades += 1
            
            current_value = cash + (btc_holdings * price)
            portfolio_values.append(current_value)
        
        final_value = portfolio_values[-1]
        total_return = (final_value - self.investment_value) / self.investment_value * 100
        
        # Calculate max drawdown
        peaks = pd.Series(portfolio_values).cummax()
        drawdowns = (pd.Series(portfolio_values) - peaks) / peaks * 100
        max_drawdown = drawdowns.min()
        
        return {
            'buy_dip_percent': buy_dip_percent,
            'sell_gain_percent': sell_gain_percent,
            'final_value': final_value,
            'total_return': total_return,
            'num_trades': num_trades,
            'total_fees_paid': total_fees_paid,
            'max_drawdown': max_drawdown
        }
    
    def optimize_parameters(self, buy_range=(1, 15), sell_range=(1, 25), step=0.5):
        """Find optimal buy/sell percentages"""
        logger.info("Starting parameter optimization...")
        
        if not self.fetch_data():
            return None
        
        # Generate parameter combinations
        buy_percentages = np.arange(buy_range[0], buy_range[1] + step, step)
        sell_percentages = np.arange(sell_range[0], sell_range[1] + step, step)
        
        results = []
        total_combinations = len(buy_percentages) * len(sell_percentages)
        
        logger.info(f"Testing {total_combinations} parameter combinations...")
        
        for i, (buy_pct, sell_pct) in enumerate(itertools.product(buy_percentages, sell_percentages)):
            # FIXED: Allow sell % to equal buy % (now includes 1%/1%)
            if sell_pct < buy_pct:  # Only skip if sell is LESS than buy
                continue
                
            result = self.run_backtest_with_params(buy_pct, sell_pct)
            if result:
                results.append(result)
            
            if (i + 1) % 50 == 0:
                logger.info(f"Processed {i + 1}/{total_combinations} combinations")
        
        # Calculate buy & hold benchmark
        initial_price = self.data['Close'].iloc[0]
        final_price = self.data['Close'].iloc[-1]
        buy_hold_return = (final_price - initial_price) / initial_price * 100
        
        # Sort by total return (descending)
        results.sort(key=lambda x: x['total_return'], reverse=True)
        
        logger.info(f"Optimization complete. Best return: {results[0]['total_return']:.2f}%")
        
        return {
            'results': results,
            'buy_hold_return': buy_hold_return,
            'total_tested': len(results),
            'data_period': f"{self.data.index[0].strftime('%Y-%m-%d')} to {self.data.index[-1].strftime('%Y-%m-%d')}"
        }
    
    def calculate_signals(self):
        """Calculate buy/sell signals based on percentage moves"""
        if self.data is None or self.data.empty:
            return
            
        self.data['Returns'] = self.data['Close'].pct_change()
        self.data['Buy_Signal'] = self.data['Returns'] <= -self.buy_dip_percent
        self.data['Sell_Signal'] = self.data['Returns'] >= self.sell_gain_percent
        
        logger.info(f"Generated {self.data['Buy_Signal'].sum()} buy signals and {self.data['Sell_Signal'].sum()} sell signals")
        
    def backtest(self):
        """Run the backtest simulation with transaction costs"""
        logger.info("Starting backtest...")
        
        if not self.fetch_data():
            logger.error("Failed to fetch data for backtest")
            return None
            
        self.calculate_signals()
        
        cash = self.investment_value
        btc_holdings = 0
        holding_btc = False
        
        portfolio_values = []
        buy_hold_values = []
        total_fees_paid = 0
        self.trades = []
        
        initial_price = self.data['Close'].iloc[0]
        btc_amount_if_bought_and_held = self.investment_value / initial_price
        
        for i, (date, row) in enumerate(self.data.iterrows()):
            price = row['Close']
            
            if row['Buy_Signal'] and not holding_btc and cash > 0:
                transaction_fee = cash * self.transaction_fee_percent
                cash_after_fees = cash - transaction_fee
                btc_bought = cash_after_fees / price
                
                btc_holdings = btc_bought
                total_fees_paid += transaction_fee
                
                self.trades.append({
                    'date': date.isoformat(),
                    'action': 'BUY',
                    'price': float(price),
                    'amount': float(btc_bought),
                    'cash_used': float(cash),
                    'fee': float(transaction_fee),
                    'cash_after_fees': float(cash_after_fees)
                })
                cash = 0
                holding_btc = True
                logger.info(f"BUY: {btc_bought:.6f} BTC at {price:.2f} (fee: £{transaction_fee:.2f}) on {date.strftime('%Y-%m-%d')}")
            
            elif row['Sell_Signal'] and holding_btc and btc_holdings > 0:
                cash_before_fees = btc_holdings * price
                transaction_fee = cash_before_fees * self.transaction_fee_percent
                cash_received = cash_before_fees - transaction_fee
                
                total_fees_paid += transaction_fee
                
                self.trades.append({
                    'date': date.isoformat(),
                    'action': 'SELL',
                    'price': float(price),
                    'amount': float(btc_holdings),
                    'cash_before_fees': float(cash_before_fees),
                    'fee': float(transaction_fee),
                    'cash_received': float(cash_received)
                })
                cash = cash_received
                btc_holdings = 0
                holding_btc = False
                logger.info(f"SELL: £{cash_received:.2f} cash at {price:.2f} (fee: £{transaction_fee:.2f}) on {date.strftime('%Y-%m-%d')}")
            
            current_strategy_value = cash + (btc_holdings * price)
            portfolio_values.append(current_strategy_value)
            
            current_buy_hold_value = btc_amount_if_bought_and_held * price
            buy_hold_values.append(current_buy_hold_value)
        
        self.data['Strategy_Value'] = portfolio_values
        self.data['Buy_Hold_Value'] = buy_hold_values
        
        final_strategy_value = portfolio_values[-1]
        final_buy_hold_value = buy_hold_values[-1]
        
        strategy_return = (final_strategy_value - self.investment_value) / self.investment_value * 100
        buy_hold_return = (final_buy_hold_value - self.investment_value) / self.investment_value * 100
        
        strategy_peaks = pd.Series(portfolio_values).cummax()
        strategy_drawdowns = (pd.Series(portfolio_values) - strategy_peaks) / strategy_peaks * 100
        max_drawdown = strategy_drawdowns.min()
        
        buy_hold_peaks = pd.Series(buy_hold_values).cummax()
        buy_hold_drawdowns = (pd.Series(buy_hold_values) - buy_hold_peaks) / buy_hold_peaks * 100
        buy_hold_max_drawdown = buy_hold_drawdowns.min()
        
        logger.info(f"Backtest complete.")
        logger.info(f"Strategy: Final value: £{final_strategy_value:.2f}, Return: {strategy_return:.2f}%")
        logger.info(f"Buy & Hold: Final value: £{final_buy_hold_value:.2f}, Return: {buy_hold_return:.2f}%")
        
        return {
            'final_value': final_strategy_value,
            'total_return': strategy_return,
            'buy_hold_return': buy_hold_return,
            'buy_hold_final_value': final_buy_hold_value,
            'num_trades': len(self.trades),
            'trades': self.trades[-10:],
            'final_position': 'BTC' if holding_btc else 'Cash',
            'total_fees_paid': total_fees_paid,
            'max_drawdown': max_drawdown,
            'buy_hold_max_drawdown': buy_hold_max_drawdown,
            'data': self.data
        }
    
    def create_chart(self, results):
        """Create enhanced interactive chart with subplots"""
        fig = make_subplots(
            rows=2, cols=1,
            subplot_titles=('BTC Price with Trading Signals', 'Portfolio Value Comparison'),
            vertical_spacing=0.1,
            row_heights=[0.6, 0.4]
        )
        
        # Price chart
        fig.add_trace(go.Scatter(
            x=self.data.index,
            y=self.data['Close'],
            name='BTC Price',
            line=dict(color='orange', width=2),
            hovertemplate='<b>Price:</b> £%{y:,.2f}<br><b>Date:</b> %{x}<extra></extra>'
        ), row=1, col=1)
        
        # Buy signals (potential)
        buy_signals = self.data[self.data['Buy_Signal']]
        if not buy_signals.empty:
            fig.add_trace(go.Scatter(
                x=buy_signals.index,
                y=buy_signals['Close'],
                mode='markers',
                name='Buy Signals',
                marker=dict(color='lightgreen', size=8, symbol='triangle-up'),
                hovertemplate='<b>Buy Signal</b><br>Price: £%{y:,.2f}<br>Date: %{x}<extra></extra>'
            ), row=1, col=1)
        
        # Sell signals (potential)
        sell_signals = self.data[self.data['Sell_Signal']]
        if not sell_signals.empty:
            fig.add_trace(go.Scatter(
                x=sell_signals.index,
                y=sell_signals['Close'],
                mode='markers',
                name='Sell Signals',
                marker=dict(color='lightcoral', size=8, symbol='triangle-down'),
                hovertemplate='<b>Sell Signal</b><br>Price: £%{y:,.2f}<br>Date: %{x}<extra></extra>'
            ), row=1, col=1)
        
        # Executed trades
        if self.trades:
            buy_trades = [t for t in self.trades if t['action'] == 'BUY']
            sell_trades = [t for t in self.trades if t['action'] == 'SELL']
            
            if buy_trades:
                fig.add_trace(go.Scatter(
                    x=[pd.to_datetime(t['date']) for t in buy_trades],
                    y=[t['price'] for t in buy_trades],
                    mode='markers',
                    name='Executed Buys',
                    marker=dict(color='darkgreen', size=15, symbol='triangle-up', 
                              line=dict(color='white', width=2)),
                    hovertemplate='<b>EXECUTED BUY</b><br>Price: £%{y:,.2f}<br>Amount: %{customdata:.6f} BTC<br>Fee: £%{meta:.2f}<extra></extra>',
                    customdata=[t['amount'] for t in buy_trades],
                    meta=[t['fee'] for t in buy_trades]
                ), row=1, col=1)
            
            if sell_trades:
                fig.add_trace(go.Scatter(
                    x=[pd.to_datetime(t['date']) for t in sell_trades],
                    y=[t['price'] for t in sell_trades],
                    mode='markers',
                    name='Executed Sells',
                    marker=dict(color='darkred', size=15, symbol='triangle-down', 
                              line=dict(color='white', width=2)),
                    hovertemplate='<b>EXECUTED SELL</b><br>Price: £%{y:,.2f}<br>Amount: %{customdata:.6f} BTC<br>Fee: £%{meta:.2f}<extra></extra>',
                    customdata=[t['amount'] for t in sell_trades],
                    meta=[t['fee'] for t in sell_trades]
                ), row=1, col=1)
        
        # Portfolio values
        fig.add_trace(go.Scatter(
            x=self.data.index,
            y=self.data['Strategy_Value'],
            name='Dip & Rip Strategy',
            line=dict(color='blue', width=3),
            hovertemplate='<b>Strategy Value:</b> £%{y:,.2f}<br><b>Date:</b> %{x}<extra></extra>'
        ), row=2, col=1)
        
        fig.add_trace(go.Scatter(
            x=self.data.index,
            y=self.data['Buy_Hold_Value'],
            name='Buy & Hold',
            line=dict(color='gray', width=2, dash='dash'),
            hovertemplate='<b>Buy & Hold Value:</b> £%{y:,.2f}<br><b>Date:</b> %{x}<extra></extra>'
        ), row=2, col=1)
        
        fig.add_hline(
            y=self.investment_value, 
            line_dash="dot", 
            line_color="black",
            annotation_text=f"Initial Investment: £{self.investment_value:,}",
            annotation_position="top right",
            row=2, col=1
        )
        
        fig.update_layout(
            title='BTC Dip & Rip Strategy Analysis',
            height=800,
            showlegend=True,
            hovermode='x unified'
        )
        
        fig.update_xaxes(title_text="Date", row=2, col=1)
        fig.update_yaxes(title_text="Price (£)", row=1, col=1)
        fig.update_yaxes(title_text="Portfolio Value (£)", row=2, col=1)
        
        return json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder)

# Utility function to fetch current BTC price
def get_current_btc_price():
    """Get current BTC/GBP price from Kraken and store as historical data"""
    try:
        url = "https://api.kraken.com/0/public/Ticker"
        params = {'pair': 'XXBTZGBP'}
        
        response = requests.get(url, params=params, timeout=5)
        response.raise_for_status()
        data = response.json()
        
        if 'error' in data and data['error']:
            # Try USD if GBP fails
            params = {'pair': 'XXBTZUSD'}
            response = requests.get(url, params=params, timeout=5)
            data = response.json()
        
        if 'result' in data and data['result']:
            for pair_name, pair_data in data['result'].items():
                if 'c' in pair_data:  # 'c' is current price
                    price = float(pair_data['c'][0])
                    currency = 'GBP' if 'GBP' in pair_name else 'USD'
                    current_time = datetime.now()
                    
                    # Store this price as historical data
                    try:
                        # Round timestamp to nearest minute to avoid too many entries
                        rounded_time = current_time.replace(second=0, microsecond=0)
                        
                        # Check if we already have data for this timestamp
                        existing = BitcoinPriceHistory.query.filter_by(timestamp=rounded_time).first()
                        
                        if not existing:
                            # Store new historical price data
                            price_record = BitcoinPriceHistory(
                                timestamp=rounded_time,
                                price_gbp=price if currency == 'GBP' else None,
                                price_usd=price if currency == 'USD' else None,
                                volume=float(pair_data.get('v', [0, 0])[1]) if 'v' in pair_data else None,  # 24h volume
                                source='kraken'
                            )
                            db.session.add(price_record)
                            db.session.commit()
                            logger.info(f"Stored historical price: {price} {currency} at {rounded_time}")
                        
                    except Exception as e:
                        logger.error(f"Failed to store historical price data: {e}")
                        db.session.rollback()
                        # Continue with returning current price even if storage fails
                    
                    return {
                        'success': True,
                        'price': price,
                        'pair': currency,
                        'timestamp': current_time.isoformat()
                    }
        
        return {'success': False, 'error': 'No price data available'}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

# Utility function to fetch historical data for data viewer
def fetch_historical_data(days=365):
    """Fetch historical BTC data for data viewer"""
    try:
        url = "https://api.kraken.com/0/public/OHLC"
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days + 30)
        since = int(start_date.timestamp())
        
        pairs_to_try = [("XXBTZGBP", "BTC-GBP"), ("XXBTZUSD", "BTC-USD")]
        
        for kraken_pair, display_name in pairs_to_try:
            try:
                params = {
                    'pair': kraken_pair,
                    'interval': 1440,
                    'since': since
                }
                
                response = requests.get(url, params=params, timeout=10)
                response.raise_for_status()
                data = response.json()
                
                if 'error' in data and data['error']:
                    continue
                
                if 'result' not in data or not data['result']:
                    continue
                
                pair_data = None
                for key in data['result'].keys():
                    if key != 'last':
                        pair_data = data['result'][key]
                        break
                
                if not pair_data or len(pair_data) < 10:
                    continue
                
                df_data = []
                for row in pair_data:
                    df_data.append({
                        'Date': datetime.fromtimestamp(int(row[0])).strftime('%Y-%m-%d'),
                        'Open': float(row[1]),
                        'High': float(row[2]),
                        'Low': float(row[3]),
                        'Close': float(row[4]),
                        'Volume': float(row[6])
                    })
                
                df = pd.DataFrame(df_data)
                df['Date'] = pd.to_datetime(df['Date'])
                df = df.sort_values('Date')
                
                return {
                    'success': True,
                    'data': df.to_dict('records'),
                    'source': display_name,
                    'total_rows': len(df)
                }
                
            except Exception as e:
                continue
        
        return {'success': False, 'error': 'All data sources failed'}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

def store_historical_price_data(hours_back=168):  # Default 7 days (168 hours)
    """Fetch and store historical BTC price data"""
    try:
        url = "https://api.kraken.com/0/public/OHLC"
        end_date = datetime.now()
        start_date = end_date - timedelta(hours=hours_back)
        since = int(start_date.timestamp())
        
        # Try GBP first, then USD
        pairs_to_try = [("XXBTZGBP", "GBP"), ("XXBTZUSD", "USD")]
        
        for kraken_pair, currency in pairs_to_try:
            try:
                params = {
                    'pair': kraken_pair,
                    'interval': 60,  # 1 hour intervals
                    'since': since
                }
                
                response = requests.get(url, params=params, timeout=10)
                response.raise_for_status()
                data = response.json()
                
                if 'error' in data and data['error']:
                    continue
                
                if 'result' not in data or not data['result']:
                    continue
                
                pair_data = None
                for key in data['result'].keys():
                    if key != 'last':
                        pair_data = data['result'][key]
                        break
                
                if not pair_data:
                    continue
                
                # Store the data
                stored_count = 0
                for row in pair_data:
                    timestamp = datetime.fromtimestamp(int(row[0]))
                    close_price = float(row[4])
                    volume = float(row[6])
                    
                    # Check if this timestamp already exists
                    existing = BitcoinPriceHistory.query.filter_by(timestamp=timestamp).first()
                    if not existing:
                        price_record = BitcoinPriceHistory(
                            timestamp=timestamp,
                            price_gbp=close_price if currency == 'GBP' else None,
                            price_usd=close_price if currency == 'USD' else None,
                            volume=volume,
                            source='kraken'
                        )
                        db.session.add(price_record)
                        stored_count += 1
                
                db.session.commit()
                return {
                    'success': True,
                    'stored_count': stored_count,
                    'currency': currency,
                    'hours_back': hours_back
                }
                
            except Exception as e:
                db.session.rollback()
                continue
        
        return {'success': False, 'error': 'All data sources failed'}
        
    except Exception as e:
        db.session.rollback()
        return {'success': False, 'error': str(e)}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def calculate_transaction_hash(transaction_data):
    """Calculate hash for duplicate detection"""
    hash_string = f"{transaction_data['date']}_{transaction_data['amount']}_{transaction_data['description']}"
    return hashlib.sha256(hash_string.encode()).hexdigest()

def extract_description_keywords(description):
    """Extract meaningful keywords from transaction description"""
    keywords = re.findall(r'\b[A-Za-z]{3,}\b', description.lower())
    return ' '.join(keywords[:5])  # Take first 5 meaningful words

def predict_category(description, existing_patterns):
    """Predict category based on learned patterns"""
    best_match = None
    best_score = 0.0
    
    description_keywords = extract_description_keywords(description)
    
    for pattern in existing_patterns:
        similarity = SequenceMatcher(None, description_keywords, pattern.description_pattern).ratio()
        score = similarity * pattern.confidence * (pattern.frequency / 10.0)  # Weight by frequency
        
        if score > best_score and score > 0.2:  # Minimum threshold (lowered for better auto-categorization)
            best_score = score
            best_match = pattern
    
    return best_match, best_score

def parse_csv_file(file_content):
    """Parse CSV file and extract transactions"""
    transactions = []
    lines = file_content.decode('utf-8').strip().split('\n')
    
    # Try to detect CSV format
    sniffer = csv.Sniffer()
    delimiter = sniffer.sniff(lines[0]).delimiter
    
    reader = csv.DictReader(lines, delimiter=delimiter)
    
    for row in reader:
        # Try common column names
        date_col = None
        amount_col = None
        desc_col = None
        
        for key in row.keys():
            key_lower = key.lower()
            if any(word in key_lower for word in ['date', 'transaction date', 'posting date']):
                date_col = key
            elif any(word in key_lower for word in ['amount', 'value', 'debit', 'credit']):
                amount_col = key
            elif any(word in key_lower for word in ['description', 'memo', 'details', 'merchant']):
                desc_col = key
        
        if date_col and amount_col and desc_col:
            try:
                # Parse date
                date_str = row[date_col]
                date_obj = pd.to_datetime(date_str).date()
                
                # Parse amount
                amount_str = str(row[amount_col]).replace(',', '').replace('£', '').replace('$', '')
                amount = float(amount_str)
                
                transactions.append({
                    'date': date_obj,
                    'amount': amount,
                    'description': row[desc_col],
                    'source_row': str(row)
                })
            except (ValueError, TypeError) as e:
                logger.warning(f"Skipping invalid row: {row} - {e}")
                continue
    
    return transactions

def parse_santander_statement(lines):
    """Parse Santander statement - only extract Money Out transactions, skip all Money In"""
    transactions = []
    
    # Find the header line to understand column structure
    header_line = None
    for line in lines:
        if 'Money In' in line and 'Money Out' in line:
            header_line = line
            break
    
    if not header_line:
        logger.warning("Could not find Money In/Money Out header in PDF")
        return []
    
    # Analyze header to find column positions
    money_out_pos = header_line.find('Money Out')
    balance_pos = header_line.find('Balance')
    
    logger.info(f"Column positions - Money Out: {money_out_pos}, Balance: {balance_pos}")
    
    for line_num, line in enumerate(lines):
        line = line.strip()
        if len(line) < 15:
            continue
            
        # Skip header and non-transaction lines
        if any(header in line.lower() for header in ['date description', 'money in', 'money out', 'balance', 'transaction date']):
            continue
            
        # Look for transaction lines starting with date
        date_match = re.match(r'(\d{1,2}\/\d{1,2}\/\d{4})', line)
        if not date_match:
            continue
            
        try:
            date_str = date_match.group(1)
            
            # Find all £ amounts in the line with their positions
            pound_matches = []
            for match in re.finditer(r'£\s*([\d,]+\.?\d*)', line):
                amount = match.group(1)
                position = match.start()
                pound_matches.append((amount, position))
            
            if len(pound_matches) < 2:
                continue
            
            # The last amount is always the balance
            balance_amount = pound_matches[-1][0]
            
            # Determine which amount is the transaction amount based on position
            transaction_amount = None
            description = ""
            
            if len(pound_matches) == 2:
                # Format: Date Description Transaction Balance
                # The first amount is the transaction (could be Money In or Money Out)
                transaction_amount = pound_matches[0][0]
                transaction_pos = pound_matches[0][1]
                
                # Extract description (between date and first £)
                date_end = date_match.end()
                description = line[date_end:transaction_pos].strip()
                
            elif len(pound_matches) == 3:
                # Format: Date Description Money_In Money_Out Balance
                money_in_amount = pound_matches[0][0]
                money_in_pos = pound_matches[0][1]
                money_out_amount = pound_matches[1][0]
                money_out_pos = pound_matches[1][1]
                
                # Extract description (between date and first £)
                date_end = date_match.end()
                description = line[date_end:money_in_pos].strip()
                
                # Check which column has a value
                money_in_val = float(money_in_amount.replace(',', ''))
                money_out_val = float(money_out_amount.replace(',', ''))
                
                if money_in_val > 0 and money_out_val == 0:
                    # This is a Money In transaction - skip it completely
                    logger.debug(f"Skipping Money In transaction: {description[:30]}... £{money_in_val}")
                    continue
                elif money_out_val > 0 and money_in_val == 0:
                    # This is a Money Out transaction - process it
                    transaction_amount = money_out_amount
                else:
                    # Both or neither have values - skip to be safe
                    continue
            else:
                continue
            
            if not transaction_amount:
                continue
                
            # Parse date
            try:
                date_obj = datetime.strptime(date_str, '%d/%m/%Y').date()
            except ValueError:
                continue
                
            # Clean description
            description = re.sub(r'\s+', ' ', description).strip()
            
            # Skip specific internal transfers that are not true expenses
            if (description == 'TRANSFER TO Shared Account' or 
                description == 'TRANSFER TO Extra Monthly Savings' or
                description == 'TRANSFER TO Savings Account' or
                description.startswith('TRANSFER TO Extra Monthly')):
                logger.debug(f"Skipping internal transfer: {description}")
                continue
            
            # Skip any Money In transactions that might slip through (based on description patterns)
            desc_lower = description.lower()
            if (desc_lower.startswith('credit from') or
                desc_lower.startswith('bank giro credit') or
                desc_lower.startswith('faster payments receipt') or
                (desc_lower == 'transfer' or desc_lower.startswith('transfer ')) and 'to' not in desc_lower):
                logger.debug(f"Skipping Money In transaction: {description}")
                continue
            
            # Convert amount to negative (expense)
            amount_val = float(transaction_amount.replace(',', ''))
            amount = -amount_val
            
            transaction = {
                'date': date_obj,
                'description': description,
                'amount': amount,
                'source_row': f"Line {line_num + 1}: {line}"
            }
            
            transactions.append(transaction)
            logger.debug(f"Added Money Out transaction: {date_obj} £{amount} '{description[:30]}...'")
            
        except (ValueError, IndexError) as e:
            logger.debug(f"Error parsing line: {line[:50]}... Error: {e}")
            continue
    
    logger.info(f"Extracted {len(transactions)} Money Out transactions from Santander statement")
    return transactions

def parse_generic_statement(lines):
    """Fallback parser for non-Santander formats"""
    # This contains the original parsing logic as backup
    # For now, return empty list - can be enhanced later if needed
    return []

def parse_pdf_file(file_content):
    """Parse PDF file using column-based extraction for Santander statements"""
    transactions = []
    
    if not PDF_SUPPORT:
        logger.error("PDF support not available. Install pdfplumber to enable PDF parsing.")
        return transactions
    
    try:
        with pdfplumber.open(io.BytesIO(file_content)) as pdf:
            full_text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    full_text += page_text + "\n"
        
        logger.info(f"PDF text length: {len(full_text)} characters")
        logger.info(f"PDF text sample: {full_text[:1000]}")
        
        lines = full_text.split('\n')
        logger.info(f"PDF has {len(lines)} lines")
        
        # Check if this is a Santander statement with Money In/Money Out columns
        is_santander_format = any('Money In' in line and 'Money Out' in line for line in lines)
        
        if is_santander_format:
            logger.info("Detected Santander format with Money In/Money Out columns")
            return parse_santander_statement(lines)
        else:
            logger.info("Using fallback parsing method")
            return parse_generic_statement(lines)
        patterns = [
            # Pattern 1: Santander format - Date Description Money Out Balance
            # Example: "22/06/2025 CARD PAYMENT TO DELIVEROO £ 29.39 £ 17.47"
            r'(\d{1,2}\/\d{1,2}\/\d{4})\s+(.+?)\s+£\s*([\d,]+\.?\d*)\s+£\s*([\d,]+\.?\d*)',
            
            # Pattern 2: Santander with Money In - Date Description Money In Balance  
            # Example: "20/06/2025 transfer £ 100.00 £ 151.80"
            r'(\d{1,2}\/\d{1,2}\/\d{4})\s+(.+?)\s+£\s*([\d,]+\.?\d*)\s+£\s*([\d,]+\.?\d*)',
            
            # Pattern 3: General UK bank format - DD/MM/YYYY description amount
            r'(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})\s+(.+?)\s+£?\s*([\d,]+\.?\d*)',
            
            # Pattern 4: Transaction spanning multiple lines (description continues)
            r'(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})\s+(.+)',
            
            # Pattern 5: Amount-heavy pattern for statements with clear money columns
            r'(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}).*?£\s*([\d,]+\.?\d*)',
        ]
        
        # Pre-process lines to handle multi-line transactions
        processed_lines = []
        in_money_in_section = False  # Track if we're in a Money In column section
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue
                
            # Check if this line starts with a date
            if re.match(r'\d{1,2}\/\d{1,2}\/\d{4}', line):
                # This is a transaction line, check if next line is a continuation
                full_line = line
                j = i + 1
                while j < len(lines) and not re.match(r'\d{1,2}\/\d{1,2}\/\d{4}', lines[j].strip()) and lines[j].strip() and 'Date Description' not in lines[j]:
                    # This line doesn't start with a date and isn't empty, it's probably a continuation
                    continuation_line = lines[j].strip()
                    # Skip lines that look like they might be amounts or balances only
                    if not re.match(r'^£\s*[\d,]+\.?\d*\s*$', continuation_line):
                        full_line += " " + continuation_line
                        logger.debug(f"Added continuation: '{continuation_line}' to line starting with '{line[:30]}...'")
                    j += 1
                processed_lines.append(full_line)
                i = j
            else:
                i += 1
        
        logger.info(f"Processed {len(processed_lines)} transaction lines from {len(lines)} total lines")
        
        for line_num, line in enumerate(processed_lines):
            line = line.strip()
            if len(line) < 15:  # Skip very short lines
                continue
                
            # Check for Money In column headers and skip transactions in those sections
            if 'money in' in line.lower():
                logger.debug(f"Skipping Money In transaction: {line[:50]}...")
                continue
                
            # Skip header lines
            if any(header in line.lower() for header in ['date description', 'money out', 'balance', 'transaction date']):
                continue
                
            # Try Santander-specific parsing first
            # New approach: Extract date first, then work backwards from the amounts to get full description
            date_match = re.match(r'(\d{1,2}\/\d{1,2}\/\d{4})', line)
            
            if date_match:
                try:
                    date_str = date_match.group(1)
                    
                    # Find all £ amounts in the line
                    pound_matches = re.findall(r'£\s*([\d,]+\.?\d*)', line)
                    
                    if len(pound_matches) >= 2:
                        # Get the last two amounts (Transaction Amount and Balance)
                        amount_str = pound_matches[-2]  # Second to last is transaction amount
                        balance_str = pound_matches[-1]  # Last is balance
                        
                        # Extract everything between the date and the last two £ symbols as description
                        # Find position after date
                        date_end = date_match.end()
                        
                        # Find position of the last two £ symbols
                        temp_line = line
                        last_pound_pos = temp_line.rfind('£')
                        temp_line = temp_line[:last_pound_pos]
                        second_last_pound_pos = temp_line.rfind('£')
                        
                        # Extract description between date and second-to-last £
                        base_description = line[date_end:second_last_pound_pos].strip()
                        
                        # Check if there's additional merchant info after the balance
                        remaining_text = line[last_pound_pos:].strip()
                        # Remove the balance amount from remaining text and extract merchant name
                        remaining_text = re.sub(r'^£\s*[\d,]+\.?\d*\s*', '', remaining_text)
                        
                        # If we find merchant info after the balance, append it to description
                        if remaining_text and not remaining_text.startswith('ON '):
                            # Extract merchant name (everything before "ON date" if present)
                            merchant_match = re.match(r'^(.+?)\s+ON\s+\d{1,2}-\d{1,2}-\d{4}', remaining_text)
                            if merchant_match:
                                merchant_name = merchant_match.group(1).strip()
                                if merchant_name:
                                    description = f"{base_description} {merchant_name}"
                                else:
                                    description = base_description
                            else:
                                # No "ON date" pattern, take everything as merchant name
                                clean_remaining = re.sub(r'\s+ON\s+\d{1,2}-\d{1,2}-\d{4}.*$', '', remaining_text).strip()
                                if clean_remaining and len(clean_remaining) > 3:
                                    description = f"{base_description} {clean_remaining}"
                                else:
                                    description = base_description
                        else:
                            description = base_description
                        
                        # Debug: Log the full extraction process
                        logger.info(f"PARSING DEBUG - Full line: '{line}'")
                        logger.info(f"PARSING DEBUG - Base description: '{base_description}'")
                        logger.info(f"PARSING DEBUG - Remaining text after balance: '{remaining_text}'")
                        logger.info(f"PARSING DEBUG - Final description: '{description}'")
                        
                        # Clean up description - remove extra whitespace but keep content
                        description = re.sub(r'\s+', ' ', description).strip()
                        
                        money_amount = amount_str
                        balance = balance_str
                        
                        logger.debug(f"Parsed line: Date='{date_str}', Description='{description}', Amount='{money_amount}', Balance='{balance}'")
                        
                        # Check if this is a Money In transaction that should be discounted
                        description_lower = description.lower()
                        
                        # Skip Money In transactions (these should be discounted/ignored)
                        if ((description_lower == 'transfer' or (description_lower.startswith('transfer ') and 'to' not in description_lower)) or
                            description_lower.startswith('deposit') or
                            description_lower.startswith('salary') or
                            description_lower.startswith('refund') or
                            'credit' in description_lower or
                            'receipt' in description_lower or
                            description_lower.startswith('bank giro credit')):
                            logger.debug(f"Skipping Money In transaction: {description[:50]}...")
                            continue
                        
                        # Now determine if it's Money Out based on description
                        amount_val = float(amount_str.replace(',', ''))
                        
                        # Default to Money Out (most transactions are expenses)
                        amount = -amount_val
                        
                        # Classify based on transaction type (check BEFORE cleaning description)
                        if (description_lower.startswith('card payment to') or 
                            description_lower.startswith('direct debit payment to') or
                            description_lower.startswith('direct debit to') or
                            'payment to' in description_lower or
                            'transfer to' in description_lower or
                            any(merchant in description_lower for merchant in ['deliveroo', 'domino', 'tesco', 'gousto', 'mcdonalds', 'google', 'vodafone', 'amazon', 'paypal', 'ebay', 'spotify', 'netflix'])):
                            # These are clearly expenses/outgoing payments
                            amount = -amount_val
                        else:
                            # For unclear cases, assume it's an expense (most bank transactions are outgoing)
                            amount = -amount_val
                            
                    else:
                        continue
                    
                    # Parse date
                    try:
                        date_obj = datetime.strptime(date_str, '%d/%m/%Y').date()
                    except ValueError:
                        continue
                    
                    # Clean description AFTER classification - only normalize spaces, keep full context
                    description = re.sub(r'\s+', ' ', description).strip()
                    
                    if abs(amount) > 0.01 and date_obj and description:  # Valid transaction
                        transactions.append({
                            'date': date_obj,
                            'amount': amount,
                            'description': description,
                            'source_row': f"Line {line_num + 1}: {line}"
                        })
                        logger.info(f"Found Santander transaction: {date_obj} £{amount} '{description}'")
                        continue  # Don't try other patterns
                        
                except (ValueError, TypeError, AttributeError) as e:
                    logger.debug(f"Error parsing Santander format {line}: {e}")
            
            # If Santander parsing failed, try other patterns
            for pattern_num, pattern in enumerate(patterns):
                matches = re.findall(pattern, line)
                if matches:
                    for match in matches:
                        try:
                            if len(match) >= 2:
                                date_str = match[0]
                                
                                # Handle different pattern structures
                                if len(match) == 3:
                                    description = match[1]
                                    amount_str = match[2]
                                elif len(match) == 4:
                                    # Pattern with description, money out, and balance
                                    description = match[1]
                                    amount_str = match[2]
                                else:
                                    amount_str = match[1] if len(match) > 1 else "0"
                                    description = match[2] if len(match) > 2 else f"Transaction from line {line_num + 1}"
                                
                                # Clean up amount string
                                amount_str = amount_str.replace('£', '').replace(',', '').strip()
                                if not amount_str or amount_str in ['', '-', '+']:
                                    continue
                                
                                # Parse date with multiple formats
                                date_obj = None
                                date_formats = ['%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y', 
                                              '%d %b %Y', '%d %B %Y', '%Y-%m-%d', '%Y/%m/%d']
                                
                                for date_format in date_formats:
                                    try:
                                        date_obj = datetime.strptime(date_str, date_format).date()
                                        break
                                    except ValueError:
                                        continue
                                
                                if not date_obj:
                                    # Try pandas for more flexible parsing
                                    try:
                                        date_obj = pd.to_datetime(date_str, dayfirst=True).date()
                                    except:
                                        continue
                                
                                # Parse amount
                                try:
                                    amount = float(amount_str)
                                except ValueError:
                                    continue
                                
                                # Clean description
                                description = description.strip()
                                if not description or len(description) < 3:
                                    description = f"Transaction on {date_obj}"
                                
                                # Validate transaction makes sense
                                if abs(amount) > 0.01 and date_obj:  # At least 1 penny
                                    transactions.append({
                                        'date': date_obj,
                                        'amount': amount,
                                        'description': description,
                                        'source_row': f"Line {line_num + 1}: {line}"
                                    })
                                    logger.info(f"Found transaction: {date_obj} £{amount} {description[:50]}")
                                    break  # Don't try other patterns for this line
                                    
                        except (ValueError, TypeError) as e:
                            logger.debug(f"Error parsing match {match}: {e}")
                            continue
                    
                    if matches:
                        break  # Don't try other patterns if we found matches
        
        # If no transactions found, try a more aggressive approach
        if not transactions:
            logger.info("No transactions found with standard patterns, trying aggressive parsing...")
            
            # Look for any line with both a date-like pattern and number
            for line_num, line in enumerate(lines):
                line = line.strip()
                
                # Find all dates in the line
                date_matches = re.findall(r'\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}', line)
                # Find all currency amounts
                amount_matches = re.findall(r'£?[\d,]+\.?\d*', line)
                
                if date_matches and amount_matches:
                    try:
                        date_str = date_matches[0]
                        amount_str = amount_matches[-1].replace('£', '').replace(',', '')  # Take last amount (often the transaction amount)
                        
                        date_obj = pd.to_datetime(date_str, dayfirst=True).date()
                        amount = float(amount_str)
                        
                        if abs(amount) > 0.01:
                            # Extract description as the text between date and amount
                            description_part = line
                            for date_match in date_matches:
                                description_part = description_part.replace(date_match, '')
                            for amount_match in amount_matches:
                                description_part = description_part.replace(amount_match, '')
                            
                            description = description_part.strip() or f"Transaction on {date_obj}"
                            
                            transactions.append({
                                'date': date_obj,
                                'amount': amount,
                                'description': description,
                                'source_row': f"Line {line_num + 1}: {line}"
                            })
                            logger.info(f"Aggressive parse found: {date_obj} £{amount} {description[:50]}")
                            
                    except (ValueError, TypeError):
                        continue
        
        logger.info(f"PDF parsing completed. Found {len(transactions)} transactions")
        
        # Debug: log first few lines of PDF for troubleshooting
        if not transactions:
            logger.warning("No transactions found. First 10 non-empty lines:")
            non_empty_lines = [line.strip() for line in lines if line.strip()]
            for i, line in enumerate(non_empty_lines[:10]):
                logger.warning(f"Line {i+1}: {line}")
    
    except Exception as e:
        logger.error(f"Error parsing PDF: {e}")
    
    return transactions

def parse_excel_file(file_content, filename):
    """Parse Excel file and extract transactions"""
    transactions = []
    
    try:
        if filename.endswith('.xls'):
            workbook = xlrd.open_workbook(file_contents=file_content)
            sheet = workbook.sheet_by_index(0)
            
            # Find header row
            headers = []
            for row_idx in range(min(5, sheet.nrows)):  # Check first 5 rows for headers
                row_values = [str(cell.value).lower() for cell in sheet.row(row_idx)]
                if any('date' in val for val in row_values):
                    headers = row_values
                    start_row = row_idx + 1
                    break
            
            if headers:
                date_col = amount_col = desc_col = None
                for i, header in enumerate(headers):
                    if 'date' in header:
                        date_col = i
                    elif any(word in header for word in ['amount', 'value', 'debit', 'credit']):
                        amount_col = i
                    elif any(word in header for word in ['description', 'memo', 'details']):
                        desc_col = i
                
                if date_col is not None and amount_col is not None and desc_col is not None:
                    for row_idx in range(start_row, sheet.nrows):
                        try:
                            row = sheet.row(row_idx)
                            date_val = row[date_col].value
                            amount_val = row[amount_col].value
                            desc_val = str(row[desc_col].value)
                            
                            # Convert Excel date to Python date
                            if isinstance(date_val, float):
                                date_obj = xlrd.xldate_as_datetime(date_val, workbook.datemode).date()
                            else:
                                date_obj = pd.to_datetime(str(date_val)).date()
                            
                            amount = float(amount_val)
                            
                            transactions.append({
                                'date': date_obj,
                                'amount': amount,
                                'description': desc_val,
                                'source_row': str([cell.value for cell in row])
                            })
                        except (ValueError, TypeError, xlrd.XLDateError):
                            continue
        
        else:  # .xlsx
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            sheet = workbook.active
            
            # Find headers
            headers = []
            for row in sheet.iter_rows(max_row=5, values_only=True):
                if any(cell and 'date' in str(cell).lower() for cell in row):
                    headers = [str(cell).lower() if cell else '' for cell in row]
                    break
            
            if headers:
                date_col = amount_col = desc_col = None
                for i, header in enumerate(headers):
                    if 'date' in header:
                        date_col = i
                    elif any(word in header for word in ['amount', 'value', 'debit', 'credit']):
                        amount_col = i
                    elif any(word in header for word in ['description', 'memo', 'details']):
                        desc_col = i
                
                if date_col is not None and amount_col is not None and desc_col is not None:
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        try:
                            if row[date_col] and row[amount_col] and row[desc_col]:
                                date_obj = pd.to_datetime(row[date_col]).date()
                                amount = float(row[amount_col])
                                description = str(row[desc_col])
                                
                                transactions.append({
                                    'date': date_obj,
                                    'amount': amount,
                                    'description': description,
                                    'source_row': str(list(row))
                                })
                        except (ValueError, TypeError):
                            continue
    
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
    
    return transactions

def get_historical_price_data(hours_back=24):
    """Get historical price data from database"""
    try:
        end_date = datetime.now()
        start_date = end_date - timedelta(hours=hours_back)
        
        price_data = BitcoinPriceHistory.query.filter(
            BitcoinPriceHistory.timestamp >= start_date,
            BitcoinPriceHistory.timestamp <= end_date
        ).order_by(BitcoinPriceHistory.timestamp.asc()).all()
        
        if not price_data:
            return {'success': False, 'error': 'No historical data available'}
        
        data_points = []
        for record in price_data:
            data_points.append({
                'timestamp': record.timestamp.isoformat(),
                'price_gbp': record.price_gbp,
                'price_usd': record.price_usd,
                'volume': record.volume
            })
        
        return {
            'success': True,
            'data': data_points,
            'count': len(data_points)
        }
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/personal-finance')
def personal_finance():
    return render_template('personal_finance.html')

@app.route('/bitcoin-tracker')
def bitcoin_tracker():
    return render_template('bitcoin_tracker.html')

@app.route('/finance-tracker')
def finance_tracker():
    return render_template('finance_tracker.html')

@app.route('/data-viewer')
def data_viewer():
    # Redirect to Bitcoin Tracker
    return redirect('/bitcoin-tracker')

@app.route('/price-monitor')
def price_monitor():
    # Redirect to Bitcoin Tracker
    return redirect('/bitcoin-tracker')

@app.route('/debug')
def debug():
    return render_template('debug.html')

@app.route('/api/current-price')
def current_price():
    """API endpoint for current BTC price"""
    price_data = get_current_btc_price()
    return jsonify(price_data)

@app.route('/api/historical-data', methods=['POST'])
def historical_data():
    """API endpoint for historical data"""
    try:
        data = request.json
        days = int(data.get('days', 365))
        
        historical = fetch_historical_data(days)
        return jsonify(historical)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dip-analysis', methods=['POST'])
def dip_analysis():
    """API endpoint for dip analysis"""
    try:
        data = request.json
        last_sell_price = float(data.get('last_sell_price', 0))
        target_dip_percent = float(data.get('target_dip_percent', 5))
        
        # Get current price
        current_price_data = get_current_btc_price()
        if not current_price_data['success']:
            return jsonify(current_price_data)
        
        current_price = current_price_data['price']
        
        # Calculate dip analysis
        if last_sell_price > 0:
            # Calculate current change from last sell price
            current_change = ((current_price - last_sell_price) / last_sell_price) * 100
            
            # Calculate target buy price (last sell price minus target dip)
            target_buy_price = last_sell_price * (1 - (target_dip_percent / 100))
            
            # Calculate how much more dip is needed
            additional_dip_needed = ((current_price - target_buy_price) / current_price) * 100
            
            # Status determination
            if current_price <= target_buy_price:
                status = "BUY_SIGNAL"
                status_text = "🟢 Target dip reached!"
            elif additional_dip_needed <= 1:
                status = "CLOSE_TO_TARGET"
                status_text = f"🟡 Very close to target (need {additional_dip_needed:.1f}% more dip)"
            elif additional_dip_needed <= 3:
                status = "APPROACHING_TARGET"
                status_text = f"🟠 Approaching target (need {additional_dip_needed:.1f}% more dip)"
            else:
                status = "FAR_FROM_TARGET"
                status_text = f"🔴 Still far from target (need {additional_dip_needed:.1f}% more dip)"
            
            return jsonify({
                'success': True,
                'current_price': current_price,
                'last_sell_price': last_sell_price,
                'target_buy_price': target_buy_price,
                'current_change_percent': current_change,
                'target_dip_percent': target_dip_percent,
                'additional_dip_needed': max(0, additional_dip_needed),
                'status': status,
                'status_text': status_text,
                'pair': current_price_data['pair'],
                'timestamp': current_price_data['timestamp']
            })
        else:
            return jsonify({
                'success': True,
                'current_price': current_price,
                'pair': current_price_data['pair'],
                'timestamp': current_price_data['timestamp'],
                'message': 'Enter your last sell price to see dip analysis'
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/backtest', methods=['POST'])
def run_backtest():
    try:
        data = request.json
        logger.info(f"Received backtest request: {data}")
        
        backtester = BTCBacktester(
            lookback_days=int(data.get('lookback_days', 365)),
            investment_value=float(data.get('investment_value', 1000)),
            buy_dip_percent=float(data.get('buy_dip_percent', 5)),
            sell_gain_percent=float(data.get('sell_gain_percent', 10)),
            transaction_fee_percent=float(data.get('transaction_fee_percent', 0.1))
        )
        
        results = backtester.backtest()
        
        if results is None:
            logger.error("Backtest returned None")
            return jsonify({'error': 'Failed to fetch market data from Kraken. Please try again in a few minutes.'}), 500
        
        chart_json = backtester.create_chart(results)
        
        return jsonify({
            'success': True,
            'results': {
                'final_value': round(results['final_value'], 2),
                'total_return': round(results['total_return'], 2),
                'buy_hold_return': round(results['buy_hold_return'], 2),
                'buy_hold_final_value': round(results['buy_hold_final_value'], 2),
                'num_trades': results['num_trades'],
                'recent_trades': results['trades'],
                'final_position': results['final_position'],
                'total_fees_paid': round(results['total_fees_paid'], 2),
                'max_drawdown': round(results['max_drawdown'], 2),
                'buy_hold_max_drawdown': round(results['buy_hold_max_drawdown'], 2)
            },
            'chart': chart_json
        })
        
    except Exception as e:
        logger.error(f"Error in backtest route: {str(e)}")
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/optimize', methods=['POST'])
def optimize_parameters():
    try:
        data = request.json
        logger.info(f"Received optimization request: {data}")
        
        backtester = BTCBacktester(
            lookback_days=int(data.get('lookback_days', 365)),
            investment_value=float(data.get('investment_value', 1000)),
            transaction_fee_percent=float(data.get('transaction_fee_percent', 0.1))
        )
        
        # Get optimization ranges
        buy_min = float(data.get('buy_min', 1))
        buy_max = float(data.get('buy_max', 15))
        sell_min = float(data.get('sell_min', 1))
        sell_max = float(data.get('sell_max', 25))
        step = float(data.get('step', 0.5))
        
        optimization_results = backtester.optimize_parameters(
            buy_range=(buy_min, buy_max),
            sell_range=(sell_min, sell_max),
            step=step
        )
        
        if optimization_results is None:
            return jsonify({'error': 'Failed to optimize parameters. Could not fetch market data.'}), 500
        
        return jsonify({
            'success': True,
            'optimization': optimization_results
        })
        
    except Exception as e:
        logger.error(f"Error in optimization route: {str(e)}")
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/test-data')
def test_data():
    try:
        url = "https://api.kraken.com/0/public/OHLC"
        params = {
            'pair': 'XXBTZUSD',
            'interval': 1440,
            'count': 5
        }
        
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        if 'error' in data and data['error']:
            return jsonify({'success': False, 'error': f'Kraken API error: {data["error"]}'})
        
        pair_data = None
        for key in data['result'].keys():
            if key != 'last':
                pair_data = data['result'][key]
                break
        
        if pair_data:
            latest_price = float(pair_data[-1][4])
            return jsonify({
                'success': True,
                'rows': len(pair_data),
                'latest_price': latest_price,
                'source': 'Kraken API'
            })
        else:
            return jsonify({'success': False, 'error': 'No price data found'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Personal Finance API Endpoints
@app.route('/api/finance/investments', methods=['GET', 'POST'])
def finance_investments():
    if request.method == 'GET':
        # Get all investments
        investments = Investment.query.order_by(Investment.created_at.desc()).all()
        
        # Get user preferences for ordering
        prefs = UserPreferences.query.first()
        investment_order = []
        if prefs and prefs.investment_order:
            investment_order = json.loads(prefs.investment_order)
        
        # Apply custom order if exists
        if investment_order:
            ordered_investments = []
            unordered_investments = []
            
            for inv_id in investment_order:
                inv = next((i for i in investments if i.id == inv_id), None)
                if inv:
                    ordered_investments.append(inv)
            
            # Add any investments not in the order
            for inv in investments:
                if inv.id not in investment_order:
                    unordered_investments.append(inv)
            
            investments = ordered_investments + unordered_investments
        
        return jsonify({
            'success': True,
            'data': [{
                'id': inv.id,
                'name': inv.name,
                'startDate': inv.start_date.isoformat(),
                'startInvestment': inv.start_investment,
                'currentValue': inv.current_value
            } for inv in investments]
        })
    
    elif request.method == 'POST':
        # Create new investment
        try:
            data = request.json
            
            investment = Investment(
                name=data['name'],
                start_date=datetime.strptime(data['startDate'], '%Y-%m-%d').date(),
                start_investment=float(data['startInvestment']),
                current_value=float(data['currentValue'])
            )
            
            db.session.add(investment)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'id': investment.id,
                    'name': investment.name,
                    'startDate': investment.start_date.isoformat(),
                    'startInvestment': investment.start_investment,
                    'currentValue': investment.current_value
                }
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/finance/investments/<int:investment_id>', methods=['PUT', 'DELETE'])
def finance_investment_detail(investment_id):
    investment = Investment.query.get_or_404(investment_id)
    
    if request.method == 'PUT':
        # Update investment
        try:
            data = request.json
            
            if 'currentValue' in data:
                investment.current_value = float(data['currentValue'])
            if 'name' in data:
                investment.name = data['name']
            if 'startInvestment' in data:
                investment.start_investment = float(data['startInvestment'])
            if 'startDate' in data:
                investment.start_date = datetime.strptime(data['startDate'], '%Y-%m-%d').date()
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'id': investment.id,
                    'name': investment.name,
                    'startDate': investment.start_date.isoformat(),
                    'startInvestment': investment.start_investment,
                    'currentValue': investment.current_value
                }
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400
    
    elif request.method == 'DELETE':
        # Delete investment and all related records
        try:
            db.session.delete(investment)
            db.session.commit()
            
            return jsonify({'success': True})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/finance/yearly-records', methods=['GET', 'POST'])
def finance_yearly_records():
    if request.method == 'GET':
        # Get all yearly records
        records = YearlyRecord.query.join(Investment).order_by(YearlyRecord.year.desc()).all()
        
        return jsonify({
            'success': True,
            'data': [{
                'id': record.id,
                'investmentId': record.investment_id,
                'investmentName': record.investment_ref.name,
                'year': record.year,
                'value': record.value,
                'notes': record.notes or '',
                'date': record.date.isoformat()
            } for record in records]
        })
    
    elif request.method == 'POST':
        # Create or update yearly record
        try:
            data = request.json
            
            # Check if record already exists
            existing_record = YearlyRecord.query.filter_by(
                investment_id=data['investmentId'],
                year=data['year']
            ).first()
            
            if existing_record:
                # Update existing record
                existing_record.value = float(data['value'])
                existing_record.notes = data.get('notes', '')
                existing_record.date = datetime.strptime(f"{data['year']}-12-31", '%Y-%m-%d').date()
                record = existing_record
            else:
                # Create new record
                record = YearlyRecord(
                    investment_id=data['investmentId'],
                    year=data['year'],
                    value=float(data['value']),
                    notes=data.get('notes', ''),
                    date=datetime.strptime(f"{data['year']}-12-31", '%Y-%m-%d').date()
                )
                db.session.add(record)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'id': record.id,
                    'investmentId': record.investment_id,
                    'investmentName': record.investment_ref.name,
                    'year': record.year,
                    'value': record.value,
                    'notes': record.notes or '',
                    'date': record.date.isoformat()
                }
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/finance/yearly-records/<int:record_id>', methods=['DELETE'])
def finance_yearly_record_detail(record_id):
    record = YearlyRecord.query.get_or_404(record_id)
    
    try:
        db.session.delete(record)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/finance/monthly-records', methods=['GET', 'POST'])
def finance_monthly_records():
    if request.method == 'GET':
        # Get all monthly records
        records = MonthlyRecord.query.join(Investment).order_by(MonthlyRecord.year.desc(), MonthlyRecord.month.desc()).all()
        
        return jsonify({
            'success': True,
            'data': [{
                'id': record.id,
                'investmentId': record.investment_id,
                'investmentName': record.investment_ref.name,
                'year': record.year,
                'month': record.month,
                'value': record.value,
                'notes': record.notes or '',
                'date': record.date.isoformat()
            } for record in records]
        })
    
    elif request.method == 'POST':
        # Create or update monthly record
        try:
            data = request.json
            
            # Check if record already exists
            existing_record = MonthlyRecord.query.filter_by(
                investment_id=data['investmentId'],
                year=data['year'],
                month=data['month']
            ).first()
            
            if existing_record:
                # Update existing record
                existing_record.value = float(data['value'])
                existing_record.notes = data.get('notes', '')
                existing_record.date = datetime.strptime(f"{data['year']}-{data['month']:02d}-01", '%Y-%m-%d').date()
                record = existing_record
            else:
                # Create new record
                record = MonthlyRecord(
                    investment_id=data['investmentId'],
                    year=data['year'],
                    month=data['month'],
                    value=float(data['value']),
                    notes=data.get('notes', ''),
                    date=datetime.strptime(f"{data['year']}-{data['month']:02d}-01", '%Y-%m-%d').date()
                )
                db.session.add(record)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'id': record.id,
                    'investmentId': record.investment_id,
                    'investmentName': record.investment_ref.name,
                    'year': record.year,
                    'month': record.month,
                    'value': record.value,
                    'notes': record.notes or '',
                    'date': record.date.isoformat()
                }
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/finance/monthly-records/<int:record_id>', methods=['DELETE'])
def finance_monthly_record_detail(record_id):
    record = MonthlyRecord.query.get_or_404(record_id)
    
    try:
        db.session.delete(record)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/finance/preferences', methods=['GET', 'POST'])
def finance_preferences():
    if request.method == 'GET':
        # Get user preferences
        prefs = UserPreferences.query.first()
        
        if prefs:
            return jsonify({
                'success': True,
                'data': {
                    'widgetOrder': json.loads(prefs.widget_order) if prefs.widget_order else ["investment", "yearly", "monthly"],
                    'investmentOrder': json.loads(prefs.investment_order) if prefs.investment_order else []
                }
            })
        else:
            return jsonify({
                'success': True,
                'data': {
                    'widgetOrder': ["investment", "yearly", "monthly"],
                    'investmentOrder': []
                }
            })
    
    elif request.method == 'POST':
        # Update user preferences
        try:
            data = request.json
            
            prefs = UserPreferences.query.first()
            if not prefs:
                prefs = UserPreferences()
                db.session.add(prefs)
            
            if 'widgetOrder' in data:
                prefs.widget_order = json.dumps(data['widgetOrder'])
            if 'investmentOrder' in data:
                prefs.investment_order = json.dumps(data['investmentOrder'])
            
            prefs.updated_at = datetime.utcnow()
            db.session.commit()
            
            return jsonify({'success': True})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/finance/backup', methods=['GET'])
def finance_backup():
    """Download database backup"""
    try:
        # Get the current database file path
        db_uri = app.config['SQLALCHEMY_DATABASE_URI']
        if not db_uri.startswith('sqlite:///'):
            return jsonify({'success': False, 'error': 'Backup only supported for SQLite databases'})
        
        # Extract database file path
        if db_uri.startswith('sqlite:////'):
            # Absolute path
            db_path = db_uri[10:]
        else:
            # Relative path
            db_path = db_uri[10:]
            
        if not os.path.exists(db_path):
            return jsonify({'success': False, 'error': 'Database file not found'})
        
        # Create timestamped filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'finance_backup_{timestamp}.db'
        
        # Create a temporary copy for download
        temp_dir = tempfile.gettempdir()
        temp_backup_path = os.path.join(temp_dir, backup_filename)
        shutil.copy2(db_path, temp_backup_path)
        
        # Send file for download
        return send_file(
            temp_backup_path,
            as_attachment=True,
            download_name=backup_filename,
            mimetype='application/octet-stream'
        )
        
    except Exception as e:
        logger.error(f"Error creating backup: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/finance/restore', methods=['POST'])
def finance_restore():
    """Restore database from uploaded file"""
    try:
        if 'backup_file' not in request.files:
            return jsonify({'success': False, 'error': 'No backup file provided'})
        
        file = request.files['backup_file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})
        
        if not file.filename.endswith('.db'):
            return jsonify({'success': False, 'error': 'Invalid file type. Please upload a .db file'})
        
        # Get current database path
        db_uri = app.config['SQLALCHEMY_DATABASE_URI']
        if not db_uri.startswith('sqlite:///'):
            return jsonify({'success': False, 'error': 'Restore only supported for SQLite databases'})
        
        if db_uri.startswith('sqlite:////'):
            db_path = db_uri[10:]
        else:
            db_path = db_uri[10:]
        
        # Create backup of current database
        if os.path.exists(db_path):
            backup_current = f"{db_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            shutil.copy2(db_path, backup_current)
            logger.info(f"Current database backed up to: {backup_current}")
        
        # Save uploaded file as new database
        file.save(db_path)
        
        # Reinitialize database connection
        db.engine.dispose()
        
        return jsonify({
            'success': True, 
            'message': 'Database restored successfully. Please refresh the page to see updated data.'
        })
        
    except Exception as e:
        logger.error(f"Error restoring backup: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

# Bitcoin Trading Tracker API Endpoints
@app.route('/api/bitcoin/trades', methods=['GET', 'POST'])
def bitcoin_trades():
    if request.method == 'GET':
        # Get all bitcoin trades
        try:
            # Use raw SQL to avoid date parsing issues
            from sqlalchemy import text
            
            result = db.session.execute(text("""
                SELECT id, status, date, type, initial_investment_gbp, 
                       btc_buy_price, btc_sell_price, profit, fee, btc_amount
                FROM bitcoin_trade
                ORDER BY date DESC
            """))
            
            trades_data = []
            for row in result:
                trades_data.append({
                    'id': row[0],
                    'status': row[1],
                    'date': row[2],
                    'type': row[3],
                    'initial_investment_gbp': row[4],
                    'btc_buy_price': row[5],
                    'btc_sell_price': row[6],
                    'profit': row[7],
                    'fee': row[8],
                    'btc_amount': row[9]
                })
            
            return jsonify({
                'success': True,
                'data': trades_data
            })
            
        except Exception as e:
            logger.error(f"Error loading bitcoin trades: {str(e)}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'POST':
        # Create new bitcoin trade
        try:
            data = request.json
            
            # Handle both old and new field names for backward compatibility
            initial_investment = float(data.get('initial_investment_gbp', data.get('gbp_investment', 0)))
            btc_buy_price = float(data.get('btc_buy_price', data.get('btc_value', 0)))
            btc_amount = initial_investment / btc_buy_price
            
            # Check if all fields are provided to determine status and calculate profit
            btc_sell_price = data.get('btc_sell_price', data.get('crypto_value'))
            fee = data.get('fee')
            
            if btc_sell_price and fee is not None:
                # All fields provided - calculate profit and set status to Closed
                btc_sell_price = float(btc_sell_price)
                fee = float(fee)
                
                # Calculate profit: (BTC amount * sell price) - initial investment - fee
                gross_return = btc_amount * btc_sell_price
                profit = gross_return - initial_investment - fee
                status = 'Closed'
            else:
                # Missing fields - set status to Open
                btc_sell_price = None
                fee = None
                profit = None
                status = 'Open'
            
            trade = BitcoinTrade(
                status=status,
                date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
                type=data.get('type', 'BTC'),
                initial_investment_gbp=initial_investment,
                btc_buy_price=btc_buy_price,
                btc_sell_price=btc_sell_price,
                profit=profit,
                fee=fee,
                btc_amount=btc_amount
            )
            
            db.session.add(trade)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'id': trade.id,
                    'status': trade.status,
                    'date': format_date_safe(trade.date),
                    'type': trade.type,
                    'initial_investment_gbp': trade.initial_investment_gbp,
                    'btc_buy_price': trade.btc_buy_price,
                    'btc_sell_price': trade.btc_sell_price,
                    'profit': trade.profit,
                    'fee': trade.fee,
                    'btc_amount': trade.btc_amount
                }
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/bitcoin/trades/<int:trade_id>', methods=['PUT', 'DELETE'])
def bitcoin_trade_detail(trade_id):
    trade = BitcoinTrade.query.get_or_404(trade_id)
    
    if request.method == 'PUT':
        # Update bitcoin trade
        try:
            data = request.json
            
            # Update fields
            if 'date' in data:
                trade.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
            if 'type' in data:
                trade.type = data['type']
            if 'initial_investment_gbp' in data:
                trade.initial_investment_gbp = float(data['initial_investment_gbp'])
            if 'btc_buy_price' in data:
                trade.btc_buy_price = float(data['btc_buy_price'])
            if 'btc_sell_price' in data:
                trade.btc_sell_price = float(data['btc_sell_price']) if data['btc_sell_price'] else None
            if 'fee' in data:
                trade.fee = float(data['fee']) if data['fee'] is not None else None
            
            # Recalculate BTC amount
            if trade.initial_investment_gbp and trade.btc_buy_price:
                trade.btc_amount = trade.initial_investment_gbp / trade.btc_buy_price
            
            # Check if all fields are complete to calculate profit and update status
            if trade.btc_sell_price is not None and trade.fee is not None and trade.btc_amount is not None:
                # Calculate profit
                gross_return = trade.btc_amount * trade.btc_sell_price
                trade.profit = gross_return - trade.initial_investment_gbp - trade.fee
                trade.status = 'Closed'
            else:
                # Not all fields complete
                trade.profit = None
                trade.status = 'Open'
            
            trade.updated_at = datetime.utcnow()
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'id': trade.id,
                    'status': trade.status,
                    'date': format_date_safe(trade.date),
                    'type': trade.type,
                    'initial_investment_gbp': trade.initial_investment_gbp,
                    'btc_buy_price': trade.btc_buy_price,
                    'btc_sell_price': trade.btc_sell_price,
                    'profit': trade.profit,
                    'fee': trade.fee,
                    'btc_amount': trade.btc_amount
                }
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400
    
    elif request.method == 'DELETE':
        # Delete bitcoin trade
        try:
            db.session.delete(trade)
            db.session.commit()
            
            return jsonify({'success': True})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/bitcoin/current-price')
def bitcoin_current_price():
    """Get current BTC price for calculations"""
    try:
        price_data = get_current_btc_price()
        if price_data:
            return jsonify({
                'success': True,
                'price': price_data['price'],
                'currency': price_data['pair']
            })
        else:
            return jsonify({'success': False, 'error': 'Could not fetch current price'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/bitcoin/historical-data')
def bitcoin_historical_data():
    """Get historical BTC price data from database"""
    try:
        hours_back = request.args.get('hours', 24, type=int)
        hours_back = min(hours_back, 8760)  # Maximum 1 year
        
        result = get_historical_price_data(hours_back)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/bitcoin/fetch-historical-data', methods=['POST'])
def bitcoin_fetch_historical_data():
    """Fetch and store historical BTC price data"""
    try:
        hours_back = request.json.get('hours', 168) if request.json else 168  # Default 7 days
        hours_back = min(hours_back, 8760)  # Maximum 1 year
        
        result = store_historical_price_data(hours_back)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dashboard/summary')
def dashboard_summary():
    """Get unified dashboard summary combining personal finance and bitcoin data"""
    try:
        # Get personal finance data
        investments = Investment.query.all()
        
        # Calculate personal finance metrics
        total_invested = sum(inv.start_investment for inv in investments)
        current_value = sum(inv.current_value for inv in investments)
        total_profit = current_value - total_invested
        profit_percentage = (total_profit / total_invested * 100) if total_invested > 0 else 0.0
        
        # Find best and worst performers
        best_performer = None
        worst_performer = None
        if investments:
            performers = []
            for inv in investments:
                profit = inv.current_value - inv.start_investment
                profit_pct = (profit / inv.start_investment * 100) if inv.start_investment > 0 else 0.0
                performers.append({'name': inv.name, 'profit_pct': profit_pct})
            
            performers.sort(key=lambda x: x['profit_pct'])
            if performers:
                worst_performer = performers[0]
                best_performer = performers[-1]
        
        # Calculate average annual return
        avg_annual_return = 0.0
        if investments:
            total_return = 0.0
            total_weight = 0.0
            for inv in investments:
                profit = inv.current_value - inv.start_investment
                years_invested = (datetime.now().date() - inv.start_date).days / 365.25
                if years_invested > 0 and inv.start_investment > 0:
                    annual_return = (profit / inv.start_investment) / years_invested * 100
                    weight = inv.start_investment
                    total_return += annual_return * weight
                    total_weight += weight
            if total_weight > 0:
                avg_annual_return = total_return / total_weight
        
        # Get Bitcoin trading data
        btc_trades = BitcoinTrade.query.all()
        
        # Calculate Bitcoin metrics
        open_positions = len([t for t in btc_trades if t.status == 'Open'])
        btc_total_profit = sum(t.profit for t in btc_trades if t.profit is not None)
        total_fees = sum(t.fee for t in btc_trades if t.fee is not None)
        
        # Get current BTC price for open position estimates
        current_btc_price = None
        price_data = get_current_btc_price()
        if price_data:
            current_btc_price = price_data['price']
            
            # Add estimated profits for open positions
            for trade in btc_trades:
                if trade.status == 'Open' and trade.btc_amount and current_btc_price:
                    estimated_value = trade.btc_amount * current_btc_price
                    estimated_profit = estimated_value - trade.initial_investment_gbp
                    btc_total_profit += estimated_profit
        
        # Find best and worst Bitcoin trades
        btc_best_trade = None
        btc_worst_trade = None
        if btc_trades:
            completed_trades = [t for t in btc_trades if t.profit is not None]
            if completed_trades:
                best_trade = max(completed_trades, key=lambda x: x.profit)
                worst_trade = min(completed_trades, key=lambda x: x.profit)
                btc_best_trade = {'profit': best_trade.profit, 'date': best_trade.date.strftime('%Y-%m-%d')}
                btc_worst_trade = {'profit': worst_trade.profit, 'date': worst_trade.date.strftime('%Y-%m-%d')}
        
        # Calculate total time invested (earliest investment)
        time_invested_months = 0
        if investments:
            earliest_date = min(inv.start_date for inv in investments)
            time_invested_months = (datetime.now().date() - earliest_date).days / 30.44
        
        # Combined portfolio value
        combined_portfolio_value = current_value + btc_total_profit
        
        # Get yearly and monthly data for historical performance
        yearly_records = YearlyRecord.query.all()
        monthly_records = MonthlyRecord.query.all()
        
        # Calculate yearly profits for historical performance
        yearly_profits = {}
        for record in yearly_records:
            year = record.year
            investment = next((inv for inv in investments if inv.id == record.investment_id), None)
            if investment:
                if year not in yearly_profits:
                    yearly_profits[year] = 0
                
                # Calculate profit for this investment in this year
                profit = record.value - investment.start_investment
                yearly_profits[year] += profit
        
        # Calculate S&P 500 comparison using actual investment period
        sp500_comparison = None
        if investments and time_invested_months > 0:
            investment_period_years = time_invested_months / 12
            
            # Use different S&P 500 rates based on time period
            if investment_period_years <= 1:
                sp500_annual_rate = 8.5  # Short term average
            elif investment_period_years <= 3:
                sp500_annual_rate = 9.2  # Medium term average
            else:
                sp500_annual_rate = 10.0  # Long term average
            
            # Calculate total returns for comparison
            my_total_return = profit_percentage
            sp500_total_return = (pow(1 + (sp500_annual_rate / 100), investment_period_years) - 1) * 100
            
            sp500_comparison = {
                'my_return': my_total_return,
                'sp500_return': sp500_total_return,
                'outperforming': my_total_return >= sp500_total_return,
                'difference': my_total_return - sp500_total_return
            }
        
        # Analyze Bitcoin trading patterns for best strategy
        best_strategy = None
        if btc_trades:
            # Analyze completed trades to find best dip/gain patterns
            completed_trades = [t for t in btc_trades if t.profit is not None and t.btc_buy_price and t.btc_sell_price]
            
            if len(completed_trades) >= 3:  # Need at least 3 trades for analysis
                strategies = []
                
                # Test different buy/sell strategies
                for buy_dip in [0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0]:
                    for sell_gain in [1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0, 7.5, 10.0]:
                        # Calculate hypothetical performance for this strategy
                        total_profit = 0
                        trade_count = 0
                        
                        for trade in completed_trades:
                            # Calculate actual dip and gain percentages
                            if trade.btc_buy_price and trade.btc_sell_price:
                                gain_pct = ((trade.btc_sell_price - trade.btc_buy_price) / trade.btc_buy_price) * 100
                                
                                # If this trade fits our strategy criteria, count it
                                if gain_pct >= sell_gain * 0.8:  # Allow some tolerance
                                    total_profit += trade.profit or 0
                                    trade_count += 1
                        
                        if trade_count > 0:
                            avg_profit = total_profit / trade_count
                            strategies.append({
                                'buy_dip': buy_dip,
                                'sell_gain': sell_gain,
                                'avg_profit': avg_profit,
                                'trade_count': trade_count,
                                'total_profit': total_profit
                            })
                
                # Find the best strategy by total profit
                if strategies:
                    best_strategy = max(strategies, key=lambda x: x['total_profit'])
        
        return jsonify({
            'success': True,
            'personal_finance': {
                'total_invested': total_invested,
                'current_value': current_value,
                'total_profit': total_profit,
                'profit_percentage': profit_percentage,
                'avg_annual_return': avg_annual_return,
                'time_invested_months': time_invested_months,
                'best_performer': best_performer,
                'worst_performer': worst_performer,
                'investment_count': len(investments),
                'sp500_comparison': sp500_comparison,
                'yearly_profits': yearly_profits
            },
            'bitcoin_trading': {
                'current_price': current_btc_price,
                'open_positions': open_positions,
                'total_profit': btc_total_profit,
                'total_fees': total_fees,
                'best_trade': btc_best_trade,
                'worst_trade': btc_worst_trade,
                'total_trades': len(btc_trades),
                'best_strategy': best_strategy
            },
            'combined': {
                'total_portfolio_value': combined_portfolio_value,
                'analysis_tools': 8  # Personal Finance (3 tabs) + Bitcoin Tracker (5 tabs)
            }
        })
        
    except Exception as e:
        logger.error(f"Dashboard summary error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Finance Tracker API Endpoints
@app.route('/api/finance-tracker/categories', methods=['GET', 'POST'])
def finance_categories():
    if request.method == 'GET':
        categories = FinanceCategory.query.order_by(FinanceCategory.name).all()
        return jsonify({
            'success': True,
            'data': [{
                'id': cat.id,
                'name': cat.name,
                'color': cat.color,
                'description': cat.description or '',
                'budget': cat.budget,
                'transaction_count': len(cat.transactions)
            } for cat in categories]
        })
    
    elif request.method == 'POST':
        try:
            data = request.json
            category = FinanceCategory(
                name=data['name'],
                color=data.get('color', '#007bff'),
                description=data.get('description', ''),
                budget=data.get('budget')
            )
            db.session.add(category)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'id': category.id,
                    'name': category.name,
                    'color': category.color,
                    'description': category.description or '',
                    'budget': category.budget,
                    'transaction_count': 0
                }
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/finance-tracker/categories/<int:category_id>', methods=['PUT', 'DELETE'])
def finance_category_detail(category_id):
    category = FinanceCategory.query.get_or_404(category_id)
    
    if request.method == 'PUT':
        try:
            data = request.json
            category.name = data.get('name', category.name)
            category.color = data.get('color', category.color)
            category.description = data.get('description', category.description)
            category.budget = data.get('budget', category.budget)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'id': category.id,
                    'name': category.name,
                    'color': category.color,
                    'description': category.description or '',
                    'transaction_count': len(category.transactions)
                }
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400
    
    elif request.method == 'DELETE':
        try:
            db.session.delete(category)
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/finance-tracker/upload', methods=['POST'])
def finance_upload():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Invalid file type. Please upload PDF files only.'}), 400
        
        filename = secure_filename(file.filename)
        file_content = file.read()
        
        # Parse PDF file
        transactions = []
        if filename.lower().endswith('.pdf'):
            transactions = parse_pdf_file(file_content)
        else:
            return jsonify({'success': False, 'error': 'Only PDF files are supported'}), 400
        
        if not transactions:
            return jsonify({'success': False, 'error': 'No valid transactions found in file'}), 400
        
        # Get existing learning patterns for auto-categorization
        patterns = FinanceCategoryLearning.query.all()
        
        # Process transactions
        processed_transactions = []
        duplicates_found = 0
        
        for trans_data in transactions:
            # Calculate hash for duplicate detection
            hash_value = calculate_transaction_hash(trans_data)
            
            # Check for existing transaction
            existing = FinanceTransaction.query.filter_by(hash_value=hash_value).first()
            if existing:
                duplicates_found += 1
                continue
            
            # Predict category
            predicted_category = None
            confidence_score = 0.0
            if patterns:
                best_pattern, score = predict_category(trans_data['description'], patterns)
                if best_pattern:
                    predicted_category = best_pattern.category_id
                    confidence_score = score
            
            # Create transaction
            transaction = FinanceTransaction(
                date=trans_data['date'],
                amount=trans_data['amount'],
                description=trans_data['description'],
                month=trans_data['date'].month,
                year=trans_data['date'].year,
                source_file=filename,
                source_row=trans_data['source_row'],
                hash_value=hash_value,
                category_id=predicted_category,
                confidence_score=confidence_score
            )
            
            db.session.add(transaction)
            processed_transactions.append({
                'date': trans_data['date'].isoformat(),
                'amount': trans_data['amount'],
                'description': trans_data['description'],
                'predicted_category': predicted_category,
                'confidence': confidence_score
            })
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'imported': len(processed_transactions),
            'duplicates_skipped': duplicates_found,
            'transactions': processed_transactions[:10]  # Show first 10 for preview
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Upload error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/finance-tracker/transactions')
def finance_transactions():
    try:
        # Get query parameters
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        category_id = request.args.get('category_id', type=int)
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        # Build query
        query = FinanceTransaction.query
        
        if year:
            query = query.filter(FinanceTransaction.year == year)
        if month:
            query = query.filter(FinanceTransaction.month == month)
        if category_id:
            query = query.filter(FinanceTransaction.category_id == category_id)
        
        # Get paginated results
        transactions = query.order_by(FinanceTransaction.date.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        return jsonify({
            'success': True,
            'data': [{
                'id': t.id,
                'date': t.date.isoformat(),
                'amount': t.amount,
                'description': t.description,
                'category_id': t.category_id,
                'category_name': t.category_ref.name if t.category_ref else None,
                'category_color': t.category_ref.color if t.category_ref else '#6c757d',
                'confidence_score': t.confidence_score,
                'is_recurring': t.is_recurring,
                'month': t.month,
                'year': t.year
            } for t in transactions.items],
            'pagination': {
                'page': transactions.page,
                'per_page': transactions.per_page,
                'total': transactions.total,
                'pages': transactions.pages
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/finance-tracker/transactions/<int:transaction_id>', methods=['PUT', 'DELETE'])
def finance_transaction_operations(transaction_id):
    try:
        transaction = FinanceTransaction.query.get_or_404(transaction_id)
        
        if request.method == 'PUT':
            # Update transaction
            data = request.json
            
            # Update fields if provided
            if 'date' in data:
                try:
                    new_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
                    transaction.date = new_date
                    transaction.year = new_date.year
                    transaction.month = new_date.month
                except ValueError:
                    return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
            
            if 'amount' in data:
                try:
                    transaction.amount = float(data['amount'])
                except ValueError:
                    return jsonify({'success': False, 'error': 'Invalid amount format'}), 400
            
            if 'description' in data:
                if len(data['description'].strip()) == 0:
                    return jsonify({'success': False, 'error': 'Description cannot be empty'}), 400
                transaction.description = data['description'].strip()
                
                # Update hash for duplicate detection
                transaction.transaction_hash = calculate_transaction_hash({
                    'date': transaction.date,
                    'amount': transaction.amount,
                    'description': transaction.description
                })
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'id': transaction.id,
                    'date': transaction.date.isoformat(),
                    'amount': transaction.amount,
                    'description': transaction.description,
                    'category_id': transaction.category_id,
                    'category_name': transaction.category_ref.name if transaction.category_ref else None,
                    'category_color': transaction.category_ref.color if transaction.category_ref else '#6c757d',
                    'confidence_score': transaction.confidence_score,
                    'month': transaction.month,
                    'year': transaction.year
                }
            })
        
        elif request.method == 'DELETE':
            # Delete transaction
            db.session.delete(transaction)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Transaction {transaction_id} deleted successfully'
            })
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/finance-tracker/transactions/delete-all', methods=['DELETE'])
def delete_all_transactions():
    try:
        # Count transactions before deletion
        transaction_count = FinanceTransaction.query.count()
        
        # Delete all transactions
        FinanceTransaction.query.delete()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'deleted_count': transaction_count,
            'message': f'Successfully deleted {transaction_count} transactions'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/finance-tracker/transactions/<int:transaction_id>/categorize', methods=['POST'])
def categorize_transaction(transaction_id):
    try:
        transaction = FinanceTransaction.query.get_or_404(transaction_id)
        data = request.json
        category_id = data.get('category_id')
        
        # Update transaction category
        transaction.category_id = category_id
        transaction.confidence_score = 1.0  # Manual categorization is 100% confident
        
        # Learn from this categorization
        if category_id:
            keywords = extract_description_keywords(transaction.description)
            
            # Check if pattern already exists
            existing_pattern = FinanceCategoryLearning.query.filter_by(
                description_pattern=keywords,
                category_id=category_id
            ).first()
            
            if existing_pattern:
                existing_pattern.frequency += 1
                existing_pattern.last_used = datetime.utcnow()
                existing_pattern.confidence = min(1.0, existing_pattern.confidence + 0.1)
            else:
                new_pattern = FinanceCategoryLearning(
                    description_pattern=keywords,
                    category_id=category_id,
                    frequency=1,
                    confidence=0.8
                )
                db.session.add(new_pattern)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': {
                'id': transaction.id,
                'category_id': transaction.category_id,
                'confidence_score': transaction.confidence_score
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/finance-tracker/summary')
def finance_summary():
    try:
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', type=int)
        
        # Build base query
        query = FinanceTransaction.query.filter(FinanceTransaction.year == year)
        if month:
            query = query.filter(FinanceTransaction.month == month)
        
        transactions = query.all()
        
        # Calculate summary metrics
        total_transactions = len(transactions)
        total_amount = sum(t.amount for t in transactions)
        income = sum(t.amount for t in transactions if t.amount > 0)
        expenses = sum(t.amount for t in transactions if t.amount < 0)
        
        # Category breakdown
        categories = {}
        for t in transactions:
            if t.category_ref:
                cat_name = t.category_ref.name
                if cat_name not in categories:
                    categories[cat_name] = {
                        'name': cat_name,
                        'color': t.category_ref.color,
                        'amount': 0,
                        'count': 0
                    }
                categories[cat_name]['amount'] += t.amount
                categories[cat_name]['count'] += 1
        
        # Monthly breakdown for the year
        monthly_data = {}
        for m in range(1, 13):
            month_transactions = [t for t in transactions if t.month == m]
            monthly_data[m] = {
                'month': m,
                'total': sum(t.amount for t in month_transactions),
                'income': sum(t.amount for t in month_transactions if t.amount > 0),
                'expenses': sum(t.amount for t in month_transactions if t.amount < 0),
                'count': len(month_transactions)
            }
        
        return jsonify({
            'success': True,
            'summary': {
                'total_transactions': total_transactions,
                'total_amount': total_amount,
                'income': income,
                'expenses': abs(expenses),
                'net': income + expenses,
                'categories': list(categories.values()),
                'monthly_data': list(monthly_data.values())
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/finance-tracker/transactions/recurring')
def finance_recurring_transactions():
    try:
        # Get all transactions marked as recurring
        recurring_transactions = FinanceTransaction.query.filter(
            FinanceTransaction.is_recurring == True
        ).all()
        
        # Calculate monthly total from recurring transactions
        recurring_total = sum(abs(t.amount) for t in recurring_transactions)
        
        return jsonify({
            'success': True,
            'recurring_total': recurring_total,
            'count': len(recurring_transactions)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/finance-tracker/transactions/recurring-detailed')
def finance_recurring_transactions_detailed():
    try:
        # Get all transactions marked as recurring with detailed information
        recurring_transactions = FinanceTransaction.query.filter(
            FinanceTransaction.is_recurring == True
        ).order_by(FinanceTransaction.date.desc()).all()
        
        # Convert to JSON format
        transactions_data = []
        for transaction in recurring_transactions:
            transactions_data.append({
                'id': transaction.id,
                'date': format_date_safe(transaction.date),
                'amount': transaction.amount,
                'description': transaction.description,
                'category_id': transaction.category_id,
                'is_recurring': transaction.is_recurring,
                'confidence_score': transaction.confidence_score,
                'created_at': format_date_safe(transaction.created_at)
            })
        
        # Calculate monthly total from recurring transactions
        recurring_total = sum(abs(t.amount) for t in recurring_transactions)
        
        return jsonify({
            'success': True,
            'transactions': transactions_data,
            'recurring_total': recurring_total,
            'count': len(transactions_data)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/finance-tracker/transactions/<int:transaction_id>/recurring', methods=['PUT'])
def update_transaction_recurring(transaction_id):
    try:
        transaction = FinanceTransaction.query.get_or_404(transaction_id)
        data = request.json
        
        transaction.is_recurring = data.get('is_recurring', False)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Recurring status updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/finance-tracker/auto-categorize', methods=['POST'])
def auto_categorize_transactions():
    """Apply auto-categorization to existing uncategorized transactions"""
    try:
        # Get all uncategorized transactions
        uncategorized = FinanceTransaction.query.filter(
            FinanceTransaction.category_id.is_(None)
        ).all()
        
        if not uncategorized:
            return jsonify({
                'success': True,
                'message': 'No uncategorized transactions found',
                'categorized_count': 0
            })
        
        # Get existing learning patterns
        patterns = FinanceCategoryLearning.query.all()
        
        if not patterns:
            return jsonify({
                'success': True, 
                'message': 'No learning patterns available. Categorize some transactions manually first.',
                'categorized_count': 0
            })
        
        categorized_count = 0
        
        for transaction in uncategorized:
            # Predict category
            best_pattern, score = predict_category(transaction.description, patterns)
            
            # Use a lower threshold for auto-categorization (0.2 instead of 0.3)
            if best_pattern and score > 0.2:
                transaction.category_id = best_pattern.category_id
                transaction.confidence_score = score
                categorized_count += 1
                
                # Update pattern usage statistics
                best_pattern.frequency += 1
                best_pattern.last_used = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully auto-categorized {categorized_count} transactions',
            'categorized_count': categorized_count,
            'total_uncategorized': len(uncategorized)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/finance-tracker/debug-pdf', methods=['POST'])
def debug_pdf():
    """Debug endpoint to examine PDF structure"""
    if not PDF_SUPPORT:
        return jsonify({'success': False, 'error': 'PDF support not available. Install pdfplumber to enable PDF parsing.'}), 400
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'success': False, 'error': 'Please upload a PDF file'}), 400
        
        file_content = file.read()
        
        # Extract text from PDF
        with pdfplumber.open(io.BytesIO(file_content)) as pdf:
            full_text = ""
            for page_num, page in enumerate(pdf.pages):
                page_text = page.extract_text()
                if page_text:
                    full_text += f"=== PAGE {page_num + 1} ===\n{page_text}\n\n"
        
        # Get lines and analyze them
        lines = full_text.split('\n')
        non_empty_lines = [line.strip() for line in lines if line.strip()]
        
        # Look for date patterns
        date_lines = []
        amount_lines = []
        
        for i, line in enumerate(non_empty_lines[:50]):  # First 50 lines
            # Check for dates
            if re.search(r'\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}', line):
                date_lines.append(f"Line {i+1}: {line}")
            
            # Check for amounts
            if re.search(r'£?[\d,]+\.?\d*', line):
                amount_lines.append(f"Line {i+1}: {line}")
        
        return jsonify({
            'success': True,
            'debug_info': {
                'total_text_length': len(full_text),
                'total_lines': len(lines),
                'non_empty_lines': len(non_empty_lines),
                'first_20_lines': non_empty_lines[:20],
                'lines_with_dates': date_lines[:10],
                'lines_with_amounts': amount_lines[:10],
                'sample_text': full_text[:2000]  # First 2000 characters
            }
        })
        
    except Exception as e:
        logger.error(f"PDF debug error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8000))
    app.run(host='0.0.0.0', port=port, debug=True)